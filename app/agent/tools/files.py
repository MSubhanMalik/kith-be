import json

from app.agent.registry import tool


@tool(description="Read and parse an Excel file (.xlsx/.xls) from a file path. Returns the data as rows and columns.")
async def read_excel(file_path: str, sheet_name: str = None, ctx=None):
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append([str(c) if c is not None else "" for c in row])
        wb.close()
        headers = rows[0] if rows else []
        data = rows[1:] if len(rows) > 1 else []
        return {"sheetName": ws.title, "headers": headers, "rows": data[:100], "totalRows": len(data)}
    except Exception as e:
        return {"error": str(e)}


@tool(description="Read and parse a CSV file from a file path. Returns the data as rows and columns.")
async def read_csv(file_path: str, ctx=None):
    try:
        import csv
        with open(file_path, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            rows = list(reader)
        headers = rows[0] if rows else []
        data = rows[1:] if len(rows) > 1 else []
        return {"headers": headers, "rows": data[:100], "totalRows": len(data)}
    except Exception as e:
        return {"error": str(e)}


@tool(description="Read a text file (.txt, .md, .json) and return its contents")
async def read_text_file(file_path: str, ctx=None):
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read(50000)
        return {"content": content, "length": len(content)}
    except Exception as e:
        return {"error": str(e)}
