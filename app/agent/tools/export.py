import os
import tempfile
from datetime import datetime

from app.agent.registry import tool
from app.services.scheduler import ScheduleService
from app.services.goal import GoalService
from app.services.task import TaskService


EXPORT_DIR = os.path.join(tempfile.gettempdir(), "kith_exports")
os.makedirs(EXPORT_DIR, exist_ok=True)


def _generate_week_xlsx(schedule_data, filename):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Weekly Schedule"

    border = Border(
        left=Side(style='thin', color='E0D8CC'), right=Side(style='thin', color='E0D8CC'),
        top=Side(style='thin', color='E0D8CC'), bottom=Side(style='thin', color='E0D8CC'),
    )
    header_fill = PatternFill(start_color='2C2417', end_color='2C2417', fill_type='solid')
    header_font = Font(bold=True, size=10, color='F5F0E8')

    ws.append(["Day", "Time", "Task", "Type", "Status"])
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    blocks = schedule_data.get("blocks", []) if schedule_data else []
    day_order = {"MON": 1, "TUE": 2, "WED": 3, "THU": 4, "FRI": 5, "SAT": 6, "SUN": 7}
    blocks.sort(key=lambda b: (day_order.get(b.get("day", ""), 8), b.get("time", {}).get("start", "")))

    for b in blocks:
        time_str = f"{b.get('time', {}).get('start', '')} - {b.get('time', {}).get('end', '')}"
        ws.append([b.get("day", ""), time_str, b.get("label", ""), b.get("type", ""), b.get("status", "")])

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    path = os.path.join(EXPORT_DIR, filename)
    wb.save(path)
    return path


def _generate_goal_xlsx(goal_data, tasks_data, filename):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = goal_data.get("label", "Goal")[:31]

    border = Border(
        left=Side(style='thin', color='E0D8CC'), right=Side(style='thin', color='E0D8CC'),
        top=Side(style='thin', color='E0D8CC'), bottom=Side(style='thin', color='E0D8CC'),
    )
    header_fill = PatternFill(start_color='2C2417', end_color='2C2417', fill_type='solid')
    header_font = Font(bold=True, size=10, color='F5F0E8')
    week_fill = PatternFill(start_color='8B7D3C', end_color='8B7D3C', fill_type='solid')
    week_font = Font(bold=True, size=11, color='FFFFFF')

    ws.merge_cells('A1:F1')
    ws['A1'] = goal_data.get("label", "")
    ws['A1'].font = Font(bold=True, size=14)

    current_week = None
    row = 3

    for t in tasks_data:
        wk = t.get("weekNumber") or 1
        if wk != current_week:
            current_week = wk
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            cell = ws.cell(row=row, column=1, value=f"WEEK {wk}")
            cell.fill = week_fill
            cell.font = week_font
            cell.alignment = Alignment(horizontal='left')
            row += 1

            for ci, h in enumerate(["Day", "Task", "Description", "Output", "Time", "Status"], 1):
                cell = ws.cell(row=row, column=ci, value=h)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
            row += 1

        ws.cell(row=row, column=1, value=t.get("dayOfWeek", ""))
        ws.cell(row=row, column=2, value=t.get("text", ""))
        ws.cell(row=row, column=3, value=t.get("description", ""))
        ws.cell(row=row, column=4, value=t.get("output", ""))
        mins = t.get("estimatedMinutes")
        ws.cell(row=row, column=5, value=f"{mins}min" if mins else "")
        ws.cell(row=row, column=6, value=t.get("status", ""))

        for ci in range(1, 7):
            ws.cell(row=row, column=ci).border = border
        row += 1

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    path = os.path.join(EXPORT_DIR, filename)
    wb.save(path)
    return path


@tool(description="Export a week's schedule as a downloadable Excel file. Returns a download URL.")
async def export_week(week_of: str, ctx=None):
    service = ScheduleService(ctx)
    data = await service.get_week(week_of)
    if not data:
        return {"error": "No schedule found for this week"}

    filename = f"kith_week_{week_of}.xlsx"
    path = _generate_week_xlsx(data, filename)
    return {"downloadUrl": f"/api/exports/download/{filename}", "filename": filename}


@tool(description="Export a goal's full plan as a downloadable Excel file. Returns a download URL.")
async def export_goal(goal_id: int, ctx=None):
    goal_service = GoalService(ctx)
    goals = await goal_service.list_goals()
    goal = next((g for g in goals if g["id"] == goal_id), None)
    if not goal:
        return {"error": "Goal not found"}

    task_service = TaskService(ctx)
    tasks = await task_service.list_tasks(goal_id)

    safe_name = goal["label"].replace(" ", "_")[:20]
    filename = f"kith_goal_{safe_name}.xlsx"
    path = _generate_goal_xlsx(goal, tasks, filename)
    return {"downloadUrl": f"/api/exports/download/{filename}", "filename": filename}
